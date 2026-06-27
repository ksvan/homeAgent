from __future__ import annotations

import hashlib
import logging
from datetime import date
from io import BytesIO
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from app.wine.models import WineBottle

logger = logging.getLogger(__name__)

# Maps lowercase source column names (and aliases) to internal field names.
_COLUMN_MAP: dict[str, str] = {
    "hylle": "shelf",
    "kategori": "category",
    "land": "country",
    "vinprodusent": "producer",
    "produsent": "producer",
    "producer": "producer",
    "vin": "name",
    "årgang": "vintage",
    "aargang": "vintage",
    "vintage": "vintage",
    "slutt drikkevindu": "drink_window_end",
    "drikkevindu slutt": "drink_window_end",
    "drink by": "drink_window_end",
    "score": "score",
    "pris innkjøp": "purchase_price_nok",
    "pris": "purchase_price_nok",
    "price": "purchase_price_nok",
    "distrikt": "region",
    "region": "region",
    "notat": "note",
    "note": "note",
    "drukket": "consumed",
    "consumed": "consumed",
    "drunk": "consumed",
}

_REQUIRED_FIELDS = {"name"}
_PREFERRED_FIELDS = {"category", "country", "producer", "vintage", "consumed"}


def parse_xlsx(
    content: bytes,
    etag: str,
    worksheet_name: str = "",
    table_name: str = "",
) -> tuple[list[dict[str, object]], list[str]]:
    """
    Parse an xlsx workbook and return (rows, warnings).

    Each row is a dict with internal field names as keys. The caller
    converts rows into WineBottle instances.
    """
    import openpyxl

    warnings: list[str] = []

    if len(content) > 10 * 1024 * 1024:
        raise ValueError("Workbook exceeds 10 MB size limit")

    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)

    ws = None
    if worksheet_name:
        ws = wb[worksheet_name] if worksheet_name in wb.sheetnames else None
        if ws is None:
            warnings.append(f"Worksheet {worksheet_name!r} not found; using first sheet")
    if ws is None:
        ws = wb.active

    if ws is None:
        raise ValueError("Workbook has no worksheets")

    rows_iter = ws.iter_rows(values_only=True)

    from typing import Any

    # Find header row: first non-empty row
    header_row: tuple[Any, ...] | None = None
    data_rows: list[tuple[Any, ...]] = []
    for row in rows_iter:
        if any(c is not None for c in row):
            if header_row is None:
                header_row = row
            else:
                data_rows.append(row)

    if header_row is None:
        raise ValueError("Could not find a header row in the workbook")

    # Build column index: raw header → internal field name
    col_index: dict[int, str] = {}
    unknown_cols: list[str] = []
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        raw = str(cell).strip()
        mapped = _COLUMN_MAP.get(raw.lower())
        if mapped:
            col_index[i] = mapped
        else:
            unknown_cols.append(raw)

    if unknown_cols:
        warnings.append(f"Unknown columns ignored: {', '.join(unknown_cols)}")

    mapped_fields = set(col_index.values())
    missing_preferred = _PREFERRED_FIELDS - mapped_fields
    if missing_preferred:
        warnings.append(f"Preferred columns missing: {', '.join(sorted(missing_preferred))}")

    if "name" not in mapped_fields:
        raise ValueError("Required column 'Vin' not found in workbook")

    parsed: list[dict[str, object]] = []
    skipped = 0

    for row_num, row in enumerate(data_rows, start=2):  # row 1 = header
        if all(c is None for c in row):
            continue  # blank row

        record: dict[str, object] = {}
        for col_i, field_name in col_index.items():
            if col_i < len(row):
                record[field_name] = row[col_i]

        name = record.get("name")
        if not name or str(name).strip() == "":
            skipped += 1
            continue

        record["name"] = str(name).strip()
        record["source_row"] = row_num
        record["source_hash"] = _row_hash(row)
        record["id"] = _make_id(row_num, etag)

        # Normalise types
        record = _coerce(record, warnings, row_num)
        parsed.append(record)

    if skipped:
        warnings.append(f"{skipped} rows skipped (missing required 'Vin' field)")

    if len(parsed) > 1000:
        warnings.append(f"Row limit exceeded; truncating to 1000 (found {len(parsed)})")
        parsed = parsed[:1000]

    wb.close()
    return parsed, warnings


def _row_hash(row: tuple[object, ...]) -> str:
    joined = "|".join("" if c is None else str(c) for c in row)
    return hashlib.md5(joined.encode()).hexdigest()[:12]


def _make_id(source_row: int, etag: str) -> str:
    return hashlib.sha1(f"{etag}:{source_row}".encode()).hexdigest()[:16]


def _coerce(record: dict[str, object], warnings: list[str], row_num: int) -> dict[str, object]:
    """Coerce raw cell values to expected Python types."""
    # vintage → int
    v = record.get("vintage")
    if v is not None:
        try:
            record["vintage"] = int(float(str(v)))
        except (ValueError, TypeError):
            warnings.append(f"Row {row_num}: invalid vintage {v!r}, ignored")
            record["vintage"] = None

    # score → float
    s = record.get("score")
    if s is not None:
        try:
            record["score"] = float(str(s))
        except (ValueError, TypeError):
            record["score"] = None

    # purchase_price_nok → float
    p = record.get("purchase_price_nok")
    if p is not None:
        try:
            record["purchase_price_nok"] = float(str(p).replace(",", "."))
        except (ValueError, TypeError):
            record["purchase_price_nok"] = None

    # drink_window_end → date
    dw = record.get("drink_window_end")
    if dw is not None:
        record["drink_window_end"] = _parse_date(dw, row_num, warnings)

    # consumed → bool: "Ja" → True, "Nei"/empty → False
    c = record.get("consumed")
    if c is None or str(c).strip() == "":
        record["consumed"] = False
    else:
        record["consumed"] = str(c).strip().lower() in ("ja", "yes", "true", "1")

    # Trim string fields
    for field in ("shelf", "category", "country", "producer", "region", "note"):
        val = record.get(field)
        if val is not None:
            stripped = str(val).strip()
            record[field] = stripped if stripped else None

    return record


def _parse_date(val: object, row_num: int, warnings: list[str]) -> date | None:
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y"):
        try:
            from datetime import datetime

            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    warnings.append(f"Row {row_num}: could not parse date {val!r}")
    return None


def rows_to_bottles(rows: list[dict[str, object]]) -> list["WineBottle"]:
    """Convert parsed row dicts to WineBottle dataclasses."""
    from app.wine.models import WineBottle

    bottles = []
    for r in rows:
        bottles.append(
            WineBottle(
                id=str(r["id"]),
                shelf=r.get("shelf"),  # type: ignore[arg-type]
                category=r.get("category"),  # type: ignore[arg-type]
                country=r.get("country"),  # type: ignore[arg-type]
                producer=r.get("producer"),  # type: ignore[arg-type]
                name=str(r["name"]),
                vintage=r.get("vintage"),  # type: ignore[arg-type]
                drink_window_end=r.get("drink_window_end"),  # type: ignore[arg-type]
                score=r.get("score"),  # type: ignore[arg-type]
                purchase_price_nok=r.get("purchase_price_nok"),  # type: ignore[arg-type]
                region=r.get("region"),  # type: ignore[arg-type]
                note=r.get("note"),  # type: ignore[arg-type]
                consumed=bool(r.get("consumed", False)),
                source_row=int(r["source_row"]),  # type: ignore[call-overload]
                source_hash=str(r["source_hash"]),
            )
        )
    return bottles
